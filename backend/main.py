from fastapi.responses import FileResponse, StreamingResponse
import io
import os
from fastapi import FastAPI, Depends, UploadFile, File, HTTPException
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse
from sqlalchemy.orm import Session
from pydantic import BaseModel
from typing import List
from database import engine, SessionLocal, Base
from models import Building
from datetime import datetime

Base.metadata.create_all(bind=engine)

app = FastAPI()

app.add_middleware(
    CORSMiddleware,
    allow_origins=["http://localhost:5173",
                   "http://localhost:5174", "http://localhost:3000"],
    allow_methods=["*"],
    allow_headers=["*"],
)

UPLOAD_DIR = "uploads"


def get_db():
    db = SessionLocal()
    try:
        yield db
    finally:
        db.close()


class BuildingIn(BaseModel):
    name: str = ""
    address: str = ""
    total_kvm: float = 0
    antal_lejemaal: int = 0
    anskaffelse: float = 0
    lejeindtægter: float = 0
    lokaleomkostninger: float = 0
    fjernvarme: float = 0
    forsikring: float = 0
    ejendomsskat: float = 0
    renovation: float = 0
    vand: float = 0
    småting: float = 0
    internet: float = 0
    ejerforening: float = 0
    administration: float = 0
    regnskabsassistance: float = 0
    vicevært: float = 0
    udvendig_vedligeholdelse: float = 0
    andet: float = 0
    omkostninger_i_alt: float = 0
    kommentar: str = ""


class BuildingOut(BuildingIn):
    id: int
    created_at: datetime | None = None

    class Config:
        from_attributes = True


@app.get("/api/health")
def health():
    return {"ok": True}


@app.post("/api/buildings", response_model=BuildingOut)
def create_building(b: BuildingIn, db: Session = Depends(get_db)):
    building = Building(**b.model_dump())
    db.add(building)
    db.commit()
    db.refresh(building)
    return building


@app.get("/api/buildings", response_model=list[BuildingOut])
def get_buildings(db: Session = Depends(get_db)):
    return db.query(Building).all()


@app.delete("/api/buildings/{building_id}")
def delete_building(building_id: int, db: Session = Depends(get_db)):
    building = db.query(Building).filter(Building.id == building_id).first()
    if not building:
        raise HTTPException(status_code=404, detail="Ejendom ikke fundet")
    db.delete(building)
    db.commit()
    return {"ok": True}


@app.post("/api/buildings/{building_id}/files")
async def upload_files(
    building_id: int,
    files: List[UploadFile] = File(...),
    db: Session = Depends(get_db),
):
    building = db.query(Building).filter(Building.id == building_id).first()
    if not building:
        raise HTTPException(status_code=404, detail="Ejendom ikke fundet")

    folder = os.path.join(UPLOAD_DIR, str(building_id))
    os.makedirs(folder, exist_ok=True)

    saved = []
    for f in files:
        path = os.path.join(folder, f.filename)
        with open(path, "wb") as out:
            out.write(await f.read())
        saved.append(f.filename)

    return {"files": saved}


@app.get("/api/buildings/{building_id}/files")
def list_files(building_id: int):
    folder = os.path.join(UPLOAD_DIR, str(building_id))
    if not os.path.exists(folder):
        return {"files": []}
    return {"files": os.listdir(folder)}


@app.get("/api/buildings/{building_id}/files/{filename}")
def download_file(building_id: int, filename: str):
    path = os.path.join(UPLOAD_DIR, str(building_id), filename)
    if not os.path.exists(path):
        raise HTTPException(status_code=404, detail="Fil ikke fundet")
    return FileResponse(path, filename=filename)


@app.delete("/api/buildings/{building_id}/files/{filename}")
def delete_file(building_id: int, filename: str):
    path = os.path.join(UPLOAD_DIR, str(building_id), filename)
    if not os.path.exists(path):
        raise HTTPException(status_code=404, detail="Fil ikke fundet")
    os.remove(path)
    return {"ok": True}


@app.put("/api/buildings/{building_id}", response_model=BuildingOut)
def update_building(building_id: int, b: BuildingIn, db: Session = Depends(get_db)):
    building = db.query(Building).filter(Building.id == building_id).first()
    if not building:
        raise HTTPException(status_code=404, detail="Ejendom ikke fundet")
    for key, value in b.model_dump().items():
        setattr(building, key, value)
    db.commit()
    db.refresh(building)
    return building


@app.post("/api/export-analyse")
def export_analyse(payload: dict):
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        import subprocess
        subprocess.check_call(
            ["pip", "install", "openpyxl", "--break-system-packages"])
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    data = payload["data"]
    if not data:
        raise HTTPException(status_code=400, detail="Ingen data")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Analyse"

    headers = list(data[0].keys())
    header_fill = PatternFill("solid", fgColor="2563EB")
    header_font = Font(bold=True, color="FFFFFF", name="Arial", size=11)
    label_font = Font(name="Arial", size=10, color="64748B")
    value_font = Font(name="Arial", size=10)
    highlight_fill = PatternFill("solid", fgColor="1E3A5F")
    thin_border = Border(
        bottom=Side(style="thin", color="334155")
    )

    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    highlight_labels = ["Omkostninger i alt",
                        "Overskud før renter", "Overskud efter renter"]

    for row_idx, row_data in enumerate(data, 2):
        is_highlight = row_data.get("", "") in highlight_labels
        for col_idx, header in enumerate(headers, 1):
            value = row_data.get(header, "")
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            if col_idx == 1:
                cell.font = Font(name="Arial", size=10,
                                 color="94A3B8", bold=is_highlight)
            else:
                cell.font = value_font
                cell.alignment = Alignment(horizontal="right")
            if is_highlight:
                cell.fill = highlight_fill
            cell.border = thin_border

    ws.column_dimensions["A"].width = 28
    for col_idx in range(2, len(headers) + 1):
        col_letter = openpyxl.utils.get_column_letter(col_idx)
        ws.column_dimensions[col_letter].width = 22

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=analyse.xlsx"}
    )
